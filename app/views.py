from app import app
from flask import Flask, render_template, request, redirect, send_from_directory, abort, flash, session, Blueprint
import os
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font
from werkzeug.utils import secure_filename
#修改excel
import pandas as pd 
from pandas import DataFrame 
import numpy as np 
import matplotlib.pyplot as plt 
from collections import Counter
#用sqlite上傳檔案
#from . import db
#用MongoDB上傳檔案
import openpyxl
import pymongo
from pymongo import MongoClient #使用mongodb
import certifi #為了解決連線到SSL的問題
import pathlib #分割副檔名
import csv
#import pandas as pd 重複了
import json
#set FLASK_ENV=development

@app.route("/") #主頁
def index():
    return render_template("public/index.html")

def allowed_excel(filename):

    #上傳的文件要有副檔名
    if not "." in filename:
        return False

    #將.從副檔名中分割出來
    ext = filename.rsplit(".", 1)[1]

    #確認副檔名和ALLOWED_EXCEL_EXTENSIONS中的一樣
    if ext.upper() in app.config["ALLOWED_EXCEL_EXTENSIONS"]:
        return True
    else:
        return False

#將檔案上傳到MongoDB
def import_files_to_mongodb(filename):
    conn = MongoClient("mongodb+srv://root:root159258@cluster0.oe4sl.mongodb.net/myFirstDatabase?retryWrites=true&w=majority", tlsCAFile=certifi.where())
    db = conn.test #選擇操作 test 資料庫
    collection = db.test2 #選擇操作 users 集合
    # test if connection success
    collection.stats  # 如果沒有error，即就連線成功

    data = pd.read_csv(app.config["EXCEL_UPLOADS"]+"/"+filename, encoding = 'utf-8-sig') #從app.config["EXCEL_UPLOADS"]裡的位置抓到上傳的檔案
    data_json = json.loads(data.to_json(orient='records'))
    collection.insert_many(data_json)                 

app.config["EXCEL_UPLOADS"] = "/app/app/static/excel" #儲存位置
app.config["ALLOWED_EXCEL_EXTENSIONS"] = ["XLSX", "XLS", "XML", "XLT", "CSV"] #允許的副檔名
app.config["SECRET_KEY"] = "OCML3BOswQEUeaxcuKHLpw" #隨機產生的SECRET_KEY，有這個才能跑flash

@app.route("/upload-excel", methods=["GET", "POST"]) #上傳excel檔
def upload_excel():

    if request.method == "POST":

        if request.files:

            excel = request.files["excel"]

            if excel.filename == "":
                flash('未選取檔案', 'warning')
                return redirect(request.url)

            if allowed_excel(excel.filename):
                filename = secure_filename(excel.filename)

                ext = filename.rsplit(".", 1)[1] #獲取檔案副檔名

                if ext == "csv":
                    excel.save(os.path.join(app.config["EXCEL_UPLOADS"], filename))
                    str_upload_path = str(app.config["EXCEL_UPLOADS"])

                    os.rename(str_upload_path + excel.filename,str_upload_path + "/ori.xlsx")

                    #excel.save(os.path.join(app.config["EXCEL_UPLOADS"], filename)) #將csv檔儲存在app.config["EXCEL_UPLOADS"]裡
                    #import_files_to_mongodb(filename)
                else:
                    excel.save(os.path.join(app.config["EXCEL_UPLOADS"], filename))
                    str_upload_path = str(app.config["EXCEL_UPLOADS"])

                    os.rename(str_upload_path + "/" + excel.filename,str_upload_path + "/" + "ori.xlsx")
                    #將excel檔案儲存在資料夾
                    #excel.save(os.path.join(app.config["EXCEL_UPLOADS"], filename))
                    #讀取使用者上傳的excel檔
                    #data_xls = pd.read_excel(app.config["EXCEL_UPLOADS"]+"/"+filename, index_col=0)

                    #由excel轉換的csv固定名稱為original.csv
                    #data_xls.to_csv(app.config["EXCEL_UPLOADS"]+"/"+'original.csv', encoding='utf-8-sig')
                    #import_file_to_mongodb('original.csv')

                    #由excel轉換的csv名稱為 原檔名.csv
                    #取出檔名(不要副檔名)
                    #filename_NoExt = os.path.splitext(filename)[0] 
                    #將 檔名.xlsx 轉換為 檔名.csv，並儲存在app.config["EXCEL_UPLOADS"]裡
                    #data_xls.to_csv(app.config["EXCEL_UPLOADS"]+"/"+filename_NoExt + '.csv', encoding='utf-8-sig') #原本是encoding='utf-8'，但下載會變成亂碼，要改encoding='utf-8-sig'
                    #import_files_to_mongodb(filename_NoExt + '.csv')

                flash('Excel saved', 'success')
                return redirect(request.url)
                #return redirect("/download/"+filename) #會下載剛剛上傳的檔案

            else:
                flash('請上傳附檔名為".xlsx .xls .xml .xlt .csv"的檔案', 'warning')
                return redirect(request.url)

    return render_template("public/upload_excel.html")

app.config["NEW_EXCEL"] = "/app/app/static/new_excel" #新excel檔的儲存位置

@app.route("/test", methods=["GET", "POST"]) #
def test():

    # 使用openpyxl建立新活頁簿wb_new
    wb_new = Workbook()
    wb_new.save(app.config["NEW_EXCEL"] + '/new_excel_test.xlsx')

    # 使用openpyxl讀取原始檔案
    wb = load_workbook(app.config["EXCEL_UPLOADS"] + '/ori.xlsx')
    ws = wb.worksheets[0]

    # 使用openpyxl讀取new_excel
    wb_new = load_workbook(app.config["NEW_EXCEL"] + '/new_excel_test.xlsx')
    ws_new = wb_new.active

    a = pd.read_excel(app.config["EXCEL_UPLOADS"] + '/ori.xlsx')
    df = pd.DataFrame(a)
    List= df['總成績'].tolist()  
    print(List)

    n=-1
    for i in List:
        n=n+1  
        df.at[n, "總成績"] = 0 
        df = DataFrame(df) 
        DataFrame(df).to_excel(app.config["EXCEL_UPLOADS"] + "/" + 'new_excel_test.xlsx', sheet_name='Sheet1', index=False, header=True)
    return redirect("/download/"+'new_excel_test.xlsx') 


#下載檔案，用from flask import send_from_directory, abort
app.config["CLIENT_EXCELS"] = "/app/app/static/excel" #要從哪裡下載

@app.route("/download/<excel_name>")
def downloadfile(excel_name):
    try:
        return send_from_directory(app.config["CLIENT_EXCELS"], path=excel_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)
#原本的程式碼return send_from_directory(app.config["CLIENT_EXCELS"], filename=excel_name, as_attachment=True)，現在filename要改成path



#將csv檔上傳到MongoDB
#conn = MongoClient("mongodb+srv://root:root159258@cluster0.oe4sl.mongodb.net/myFirstDatabase?retryWrites=true&w=majority", tlsCAFile=certifi.where())
#db = conn.test #選擇操作 test 資料庫
#collection = db.users #選擇操作 users 集合

# test if connection success
#collection.stats  # 如果沒有error，即就連線成功

#df = pd.read_csv("test.csv") #選擇要操作的csv檔
#data = df.to_dict(orient = "records")
#collection.insert_many(data) #新增資料